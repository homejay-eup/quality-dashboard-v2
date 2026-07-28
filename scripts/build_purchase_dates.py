"""
一次性轉換腳本：讀取桌面「2026進貨0728_所有條碼.XLSX」，篩選「單據名稱＝產品進貨單」，
展開序號範圍（數量>1 的批次進貨列），依條碼去重取最舊「單據日期」，
輸出 js/data/purchase_dates.json 供 sheets.js 靜態載入。

用法：python scripts/build_purchase_dates.py
"""
import json
import re
import sys
from datetime import datetime

import openpyxl

SRC = r"C:\Users\EupUser\Desktop\品質分析\2026進貨0728_所有條碼.XLSX"
OUT = r"C:\Users\EupUser\Desktop\品質分析\.claude\worktrees\equipment-quality-purchase-date-01fbe0\js\data\purchase_dates.json"

TARGET_DOC_NAME = "產品進貨單"


def parse_date(s):
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime(y, mo, d)
    except ValueError:
        return None


def expand_range(start, end, qty):
    """展開序號範圍成個別條碼清單。start==end 或無法展開時回傳 [start]。"""
    if not start:
        return []
    if start == end or not end or not qty or qty <= 1:
        return [start]

    # 找出 start/end 共同前綴之後、長度相同的數字尾碼，嘗試不同尾碼寬度
    max_len = min(len(start), len(end))
    for k in range(1, max_len + 1):
        s_prefix, s_suffix = start[:-k], start[-k:]
        e_prefix, e_suffix = end[:-k], end[-k:]
        if s_prefix != e_prefix:
            continue
        if not (s_suffix.isdigit() and e_suffix.isdigit()):
            continue
        s_num, e_num = int(s_suffix), int(e_suffix)
        if e_num - s_num == qty - 1:
            width = len(s_suffix)
            return [f"{s_prefix}{str(s_num + i).zfill(width)}" for i in range(qty)]

    # 找不到吻合的展開規則，退回只用起始編號
    return [start]


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    result = {}
    total_rows = 0
    total_barcodes = 0
    expand_used = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        doc_name = row[1]
        if doc_name != TARGET_DOC_NAME:
            continue
        total_rows += 1

        doc_date_str = row[2]
        start, end, qty = row[8], row[9], row[10]

        d = parse_date(doc_date_str)
        if not d:
            continue

        barcodes = expand_range(start, end, qty)
        if start != end and len(barcodes) > 1:
            expand_used += 1

        for bc in barcodes:
            if not bc:
                continue
            total_barcodes += 1
            prev = result.get(bc)
            if prev is None or d < prev:
                result[bc] = d

    out = {bc: d.strftime("%Y-%m-%d") for bc, d in result.items()}

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    print(f"產品進貨單 rows: {total_rows}")
    print(f"barcode entries written (post expand+dedupe): {len(out)}")
    print(f"expand_range 觸發次數（qty>1 且展開成功）: {expand_used}")
    print(f"total barcode occurrences seen (含重複): {total_barcodes}")
    print(f"輸出檔案: {OUT}")


if __name__ == "__main__":
    main()
